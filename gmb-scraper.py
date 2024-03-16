import os
import re
import time
import pyautogui
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from dotenv import find_dotenv, load_dotenv

load_dotenv(find_dotenv())

##################################
## START - DOM ATTRIBUTE
##################################
get_map_search_box = '//*[(@id = "searchboxinput")]'
get_business_from_sidebar = "hfpxzc"

get_business_name = '//*[contains(concat( " ", @class, " " ), concat( " ", "lfPIob", " " ))]'
get_business_address = '//*[contains(concat( " ", @class, " " ), concat( " ", "AG25L", " " )) and (((count(preceding-sibling::*) + 1) = 3) and parent::*)]//*[contains(concat( " ", @class, " " ), concat( " ", "kR99db", " " ))]'
get_business_website = '//*[contains(concat( " ", @class, " " ), concat( " ", "ITvuef", " " ))]//*[contains(concat( " ", @class, " " ), concat( " ", "kR99db", " " ))]'
# get_business_phone = '//*[contains(concat( " ", @class, " " ), concat( " ", "AG25L", " " )) and (((count(preceding-sibling::*) + 1) = 7) and parent::*)]//*[contains(concat( " ", @class, " " ), concat( " ", "kR99db", " " ))]'
get_business_phone = '//*[contains(concat( " ", @class, " " ), concat( " ", "AG25L", " " )) and (((count(preceding-sibling::*) + 1) = 8) and parent::*)]//*[contains(concat( " ", @class, " " ), concat( " ", "kR99db", " " ))]'
                      
##################################
## END - DOM ATTRIBUTE
##################################


#####################################
## START - SAVE BUSINESS INFORMATION TO EXCEL
#####################################
def append_to_excel(data):
    # Load the existing workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()

    ws = wb.active

    # Check if the business name or website already exists in the Excel file
    existing_businesses = set()
    column_names = [cell.value for cell in ws[1]]  # Assuming the column names are in the first row
    business_name_index = column_names.index("Business Name")
    website_index = column_names.index("Website")

    for row in ws.iter_rows(min_row=2, values_only=True):  # Start from the second row
        existing_businesses.add(row[business_name_index])
        existing_businesses.add(row[website_index])

    if data[business_name_index] not in existing_businesses and data[website_index] not in existing_businesses:
        ws.append(data)
        wb.save(excel_file)
        print("Data appended to the Excel file.")
    else:
        print("Business name or website already exists. Skipping data.")
#####################################
## START - SAVE BUSINESS INFORMATION TO EXCEL
#####################################



#####################################
## START - GET BUSINESS INFORMATION
#####################################
def get_business_information(driver, business):
    business.click()

    time.sleep(2)
    # CLICK THE BUSINESS TO GET DETAILS
    try:
        business_profile_name = driver.find_element(By.XPATH, get_business_name)
        if business_profile_name:
            business_name = business_profile_name.text.strip()
        else:
            # Element not found, set default value
            business_name = 'N/A'
    except NoSuchElementException:
        print("Element not found. Setting default value.")
        business_name = 'N/A'

    time.sleep(2)
    try:
        business_profile_address = driver.find_element(By.XPATH, get_business_address)
        if business_profile_address:
            print('Business profile Address --->', business_profile_address.text.strip())
            business_address = business_profile_address.text.strip()
        else:
            # Element not found, set default value
            business_address = 'N/A'
    except NoSuchElementException:
        print("Element not found. Setting default value.")
        business_address = 'N/A'


    time.sleep(2)
    try:
        business_profile_website = driver.find_element(By.XPATH, get_business_website)
        if business_profile_website:
            print('Business profile Website --->', business_profile_website.text.strip())
            business_website = business_profile_website.text.strip()
        else:
            # Element not found, set default value
            business_website = 'N/A'
    except NoSuchElementException:
        print("Element not found. Setting default value.")
        business_website = 'N/A'


    time.sleep(2)
    try:
        # phone_pattern = "^\\+?\\d{1,4}?[-.\\s]?\\(?\\d{1,3}?\\)?[-.\\s]?\\d{1,4}[-.\\s]?\\d{1,4}[-.\\s]?\\d{1,9}$"
        phone_pattern = "^(\+?\d{1,4}[-. ]?)?(\(\d{3}\)|\d{3})[-. ]?\d{3}[-. ]?\d{4}$"

        # Find all details
        details = driver.find_elements(By.CLASS_NAME, "kR99db")
        
        # Initialize variables
        business_phone = 'N/A'
        
        # Search for phone number within details
        for detail in details:
            text = detail.text.strip()
            print('inside text: ', text)
            if re.search(phone_pattern, text):
                business_phone = text
                break  # Stop searching once phone number is found
        
        print('Business profile Phone --->', business_phone)
        
    except NoSuchElementException:
        print("Element not found. Setting default value.")
        business_phone = 'N/A'



    # Append data to the Excel file
    data = (business_name, 'N/A', business_phone, business_website, 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', business_address)

    append_to_excel(data)

    time.sleep(5)

#####################################
## END - GET BUSINESS INFORMATION
#####################################




#####################################
## START - SCROLL FOR EACH LOCATION
#####################################
def scroll_to_get_business_profile(driver, location):
    map_search_box = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, get_map_search_box)))
    map_search_box.clear()

    search_field_element = f'"{search_keyword}" + "{location}" + "{search_country}"'

    map_search_box.send_keys(search_field_element)
    map_search_box.send_keys(Keys.ENTER)
    time.sleep(10)

    # Scroll to collect business profiles
    for _ in range(200):  # Adjust the range based on the number of scrolls needed
        pyautogui.scroll(-1000)  # Scroll up 1000 pixels
        time.sleep(2)  # Wait for the page to load after scrolling

    # CLICK THE BUSINESS TO GET DETAILS
    get_businesses_profile = driver.find_elements(By.CLASS_NAME, get_business_from_sidebar)
    # business_anchors = [a.get_attribute('href') for a in get_businesses_profile]
    for business in get_businesses_profile:
        get_business_information(driver, business)
        time.sleep(5)
#####################################
## END - SCROLL FOR EACH LOCATION
#####################################



#####################################
## START - MAIN TASK
#####################################
def main():
    # Configure driver
    service = Service(executable_path=os.getenv('CHROME_DRIVER'))
    options = webdriver.ChromeOptions()
    prefs = {
        "profile.default_content_setting_values.notifications": 2
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=service, options=options)

    # Open the website and login
    driver.get("https://google.com/maps/")
    
    for location in search_locations:
        print('---- STARTING BUSINESS LOOP -----')
        scroll_to_get_business_profile(driver, location)
        print('---- END BUSINESS LOOP -----')
        time.sleep(5)

    # Close the driver
    driver.quit()

#####################################
## END - MAIN TASK
#####################################


if __name__ == "__main__":
    excel_file = "gmb_business_lead.xlsx"

    # Set parameters
    search_country = "USA"
    search_keyword = "Gymcenter"
    search_locations = [
        "Phoenix",
        "Charlotte",
        "Las Vegas",
        "Jacksonville"
    ]

    # EXCEL COLUMN NAME
    data = [
        (
            "Business Name", "Email", "Phone", "Website", 
            "Facebook", "Twitter", "Instagram", "Youtube", "Linkedin",
            "Location"
        )
    ]

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_file)

    main()